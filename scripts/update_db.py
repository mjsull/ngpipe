import os
from urllib.request import urlopen
import json
from openpyxl import load_workbook
import io
import datetime


db = snakemake.params.db
if db == "ngstar":
    scheme = "67"
elif db == "ngmast":
    scheme = "71"
elif db == "mlst":
    scheme = "1"
elif db == "rplf":
    scheme = "42"

params_mlst_dir = snakemake.params.mlst_dir

dbdir = os.path.join(params_mlst_dir, 'db', 'pubmlst', db)
version_log = os.path.join(dbdir, "version.log")
if not os.path.exists(dbdir):
    os.makedirs(dbdir)
if os.path.exists(version_log):
    with open(version_log) as f:
        version = f.readline().rstrip()
else:
    version = None




if snakemake.params.update_db != "no":
    with urlopen("https://rest.pubmlst.org/db/pubmlst_neisseria_seqdef/schemes/" + scheme) as response:
        response_content = response.read().decode('utf-8')
    scheme_info = json.loads(response_content)
    online_version = scheme_info["last_added"]



if snakemake.params.update_db == "no":
    with open(snakemake.output.log, 'w') as o:
        o.write("Database update not requested.\n")
elif version != online_version and db != "ngstar":
    profiles = scheme_info["profiles_csv"]
    response = urlopen(profiles)
    data = response.read()
    profiles_local = os.path.join(dbdir, db + '.txt')
    with open(profiles_local, 'wb') as o:
        o.write(data)
    for i in scheme_info['loci']:
        loci = i.split('/')[-1]
        response = urlopen(i)
        loci_info = json.loads(response.read().decode('utf-8'))
        loci_fasta = loci_info["alleles_fasta"]
        response = urlopen(loci_fasta)
        data = response.read()
        loci_local = os.path.join(dbdir, loci + '.tfa')
        loci_local = loci_local.replace("'", "")
        loci_local = loci_local.replace("NG-MAST_", "")
        with open(loci_local, 'wb') as o:
            o.write(data)


    with open(version_log, 'w') as o:
        o.write(online_version + "\n")
    with open(snakemake.output.log, 'w') as o:
        o.write("Database updated to version uploaded on " + online_version + "\n")
elif db == "ngstar":
    locistar = ['penA', 'mtrR', 'porB', 'ponA', 'gyrA', 'parC', '23S']
    ngstar_loci = 'https://ngstar.canada.ca/alleles/download?lang=en&loci_name='
    ngstar_profiles = 'https://ngstar.canada.ca/sequence_types/download?lang=en'
    ngstar_meta = 'https://ngstar.canada.ca/alleles/download_metadata?lang=en&loci_name='
    # Download allele fasta files
    for loci in locistar:
        response = urlopen(ngstar_loci + loci)
        data = response.read().decode()
        loci_local = os.path.join(dbdir, loci + '.tfa')
        with open(loci_local, 'w') as o:
            if loci == "penA":
                for line in data.split("\n"):
                    if line.startswith(">"):
                        pena_loci = line.split("_")[1]
                        o.write(">penA_{:.3f}\n".format(float(pena_loci)).replace(".", ""))
                    else:
                        o.write("{}\n".format(line))
            else:
                o.write(data.replace(".0\n", "\n"))
                o.write("\n")
        response = urlopen(ngstar_meta + loci)
        bytes_in = io.BytesIO(response.read())
        wb = load_workbook(bytes_in)
        ws = wb.active
        comment_local = os.path.join(dbdir, loci + '.tfa.comments')
        with open(comment_local, 'w') as o:
            for row in ws.iter_rows(values_only=True):
                o.write("\t".join(map(str, row)).replace("\n", " ") + "\n")
    response = urlopen(ngstar_profiles)
    bytes_in = io.BytesIO(response.read())
    wb = load_workbook(bytes_in)
    profiles_local = os.path.join(dbdir, db + '.txt')
    ws = wb.active
    with open(profiles_local, 'w') as o:
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Sequence Type":
                o.write("ST")
                for j in row[1:]:
                    o.write("\t{}".format(j))
                o.write("\n")
            else:
                o.write(str(row[0]))
                for j in row[1:]:
                    if "." in str(j):
                        o.write("\t{:.3f}".format(j).replace(".", ""))
                    else:
                        o.write("\t{}".format(j))
                o.write("\n")
    with open(snakemake.output.log, 'w') as o:
        o.write("Using ngstar donwloaded {}\n".format(datetime.datetime.now()))
                

else:
    with open(snakemake.output.log, 'w') as o:
        o.write("Database already using version uploaded on " + online_version + "\n")
